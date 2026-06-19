import io
from datetime import date


def generate_analytics_report(
    financial: dict,
    attendance: dict,
    performance: dict,
    period_days: int,
) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # ---- Sheet 1: Cover ----
    ws = wb.active
    ws.title = "Зведення"
    ws.sheet_properties.tabColor = "2E75B6"
    ws.column_dimensions["A"].width = 50

    gen_date = date.today().strftime("%d.%m.%Y")
    ws["A1"] = "AleksSchool — Аналітичний звіт"
    ws["A1"].font = Font(size=16, bold=True, color="1F4E79")
    ws["A2"] = f"Дата генерації: {gen_date}"
    ws["A3"] = f"Аналізований період: останні {period_days} днів"

    ws["A5"] = "💰 Фінанси"
    ws["A5"].font = Font(bold=True)
    ws["A6"] = f"Підтверджено: {financial.get('total_confirmed', 0):,.0f} грн  ({financial.get('confirmed_count', 0)} оплат)"
    ws["A7"] = f"Боржників: {financial.get('debtors_count', 0)} учнів"

    by_group = attendance.get("by_group", [])
    avg_att = round(sum(g["percent"] for g in by_group) / len(by_group)) if by_group else 0
    ws["A9"] = "📊 Відвідуваність"
    ws["A9"].font = Font(bold=True)
    ws["A10"] = f"Груп проаналізовано: {len(by_group)}"
    ws["A11"] = f"Середня відвідуваність: {avg_att}%"

    hw = performance.get("homework", {})
    qz = performance.get("quizzes", {})
    ws["A13"] = "📚 Успішність"
    ws["A13"].font = Font(bold=True)
    ws["A14"] = f"ДЗ здано: {hw.get('completion_rate', 0)}%  ({hw.get('submitted_count', 0)}/{hw.get('assigned_count', 0)})"
    ws["A15"] = f"Тести: {qz.get('avg_score_pct', 0)}% середній результат"

    # ---- Sheet 2: Financial ----
    wf = wb.create_sheet("Фінанси")
    wf.sheet_properties.tabColor = "2E75B6"

    _header(wf, 1, ["Показник", "Значення"], "1F4E79")
    fin_rows = [
        ("Підтверджено (грн)", financial.get("total_confirmed", 0)),
        ("Очікують підтвердження (грн)", financial.get("total_pending", 0)),
        ("Відхилено (грн)", financial.get("total_rejected", 0)),
        ("К-сть підтверджених оплат", financial.get("confirmed_count", 0)),
        ("К-сть очікуючих оплат", financial.get("pending_count", 0)),
        ("К-сть боржників", financial.get("debtors_count", 0)),
        ("Орієнтовна сума боргу (грн)", financial.get("debtors_total_estimate", 0)),
    ]
    for i, (label, value) in enumerate(fin_rows, 2):
        wf.cell(row=i, column=1, value=label)
        cell_val = round(float(value), 2) if isinstance(value, float) else value
        wf.cell(row=i, column=2, value=cell_val)

    row = len(fin_rows) + 3
    wf.cell(row=row, column=1, value="Виручка по місяцях (підтверджені оплати)")
    wf.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    _header(wf, row, ["Місяць", "Виручка (грн)", "К-сть оплат"], "1F4E79")
    for m in financial.get("monthly_revenue", []):
        row += 1
        wf.cell(row=row, column=1, value=m["month"])
        wf.cell(row=row, column=2, value=round(m["revenue"], 2))
        wf.cell(row=row, column=3, value=m["count"])

    row += 2
    wf.cell(row=row, column=1, value="Список боржників")
    wf.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    _header(wf, row, ["Ім'я учня", "Групи", "Остання підтверджена оплата"], "1F4E79")
    red_fill = PatternFill("solid", fgColor="FCE4D6")
    for d in financial.get("debtors", []):
        row += 1
        wf.cell(row=row, column=1, value=d.get("full_name") or "—")
        wf.cell(row=row, column=2, value=", ".join(d.get("group_names", [])))
        last = d.get("last_payment_at")
        wf.cell(row=row, column=3, value=(last[:10] if last else "Ніколи"))
        for col in range(1, 4):
            wf.cell(row=row, column=col).fill = red_fill

    _auto_width(wf)

    # ---- Sheet 3: Attendance ----
    wa = wb.create_sheet("Відвідуваність")
    wa.sheet_properties.tabColor = "375623"

    _header(wa, 1, ["Група", "Занять", "Учнів", "Відвідано", "Можливо", "%"], "1E5631")
    for i, g in enumerate(attendance.get("by_group", []), 2):
        wa.cell(row=i, column=1, value=g["group_name"])
        wa.cell(row=i, column=2, value=g["total_lessons"])
        wa.cell(row=i, column=3, value=g["student_count"])
        wa.cell(row=i, column=4, value=g["present_count"])
        wa.cell(row=i, column=5, value=g["possible_count"])
        pct_cell = wa.cell(row=i, column=6, value=g["percent"])
        _color_pct(pct_cell, g["percent"])

    row = len(attendance.get("by_group", [])) + 3
    _header(wa, row, ["Учень", "Група", "Занять", "Присутній", "Запізнився", "Відсутній", "Поважна", "%"], "1E5631")
    for s in attendance.get("by_student", []):
        row += 1
        wa.cell(row=row, column=1, value=s.get("full_name") or "—")
        wa.cell(row=row, column=2, value=s.get("group_name") or "—")
        wa.cell(row=row, column=3, value=s["total"])
        wa.cell(row=row, column=4, value=s["present"])
        wa.cell(row=row, column=5, value=s["late"])
        wa.cell(row=row, column=6, value=s["absent"])
        wa.cell(row=row, column=7, value=s["excused"])
        pct_cell = wa.cell(row=row, column=8, value=s["percent"])
        _color_pct(pct_cell, s["percent"])

    _auto_width(wa)

    # ---- Sheet 4: Performance ----
    wp = wb.create_sheet("Успішність")
    wp.sheet_properties.tabColor = "4A235A"

    wp["A1"] = "Домашні завдання"
    wp["A1"].font = Font(bold=True, color="4A235A")
    wp["A2"] = f"Призначено: {hw.get('assigned_count', 0)}"
    wp["A3"] = f"Здано: {hw.get('submitted_count', 0)} ({hw.get('completion_rate', 0)}%)"
    wp["A5"] = "Тести"
    wp["A5"].font = Font(bold=True, color="4A235A")
    wp["A6"] = f"Спроб всього: {qz.get('attempts_count', 0)}"
    wp["A7"] = f"Завершено: {qz.get('completed_count', 0)}"
    wp["A8"] = f"Середній результат: {qz.get('avg_score_pct', 0)}%"

    row = 10
    _header(wp, row, ["Учень", "Група", "ДЗ %", "Тести %", "Загальний бал"], "4A235A")
    for s in performance.get("student_ranking", []):
        row += 1
        wp.cell(row=row, column=1, value=s.get("full_name") or "—")
        wp.cell(row=row, column=2, value=s.get("group_name") or "—")
        wp.cell(row=row, column=3, value=s["hw_completion_pct"])
        wp.cell(row=row, column=4, value=s["quiz_avg_pct"])
        score_cell = wp.cell(row=row, column=5, value=s["combined_score"])
        _color_score(score_cell, s["combined_score"])

    _auto_width(wp)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header(ws, row: int, titles: list[str], color: str) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF")
    for col, title in enumerate(titles, 1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _color_pct(cell, pct: int) -> None:
    from openpyxl.styles import PatternFill
    if pct >= 80:
        cell.fill = PatternFill("solid", fgColor="E2EFDA")
    elif pct >= 60:
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
    else:
        cell.fill = PatternFill("solid", fgColor="FCE4D6")


def _color_score(cell, score: float) -> None:
    from openpyxl.styles import PatternFill
    if score >= 8.0:
        cell.fill = PatternFill("solid", fgColor="E2EFDA")
    elif score >= 5.0:
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
    else:
        cell.fill = PatternFill("solid", fgColor="FCE4D6")


def _auto_width(ws) -> None:
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=8)
        ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 40))

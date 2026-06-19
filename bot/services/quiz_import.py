import io


def generate_quiz_template() -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quiz"

    white_bold = Font(bold=True, color="FFFFFF")
    blue_fill = PatternFill("solid", fgColor="4472C4")
    green_fill = PatternFill("solid", fgColor="70AD47")

    # Row 1: quiz metadata headers
    for col, h in enumerate(["title", "description", "time_limit_min", "shuffle_questions"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_bold
        cell.fill = blue_fill

    # Row 2: example values
    ws.cell(row=2, column=1, value="Назва тесту")
    ws.cell(row=2, column=2, value="Опис (необов'язково)")
    ws.cell(row=2, column=3, value=30)
    ws.cell(row=2, column=4, value="FALSE")

    # Row 4: question headers
    for col, h in enumerate(["question", "type", "option1", "option2", "option3", "option4", "correct"], 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = white_bold
        cell.fill = green_fill

    # Row 5: single-choice example
    ws.cell(row=5, column=1, value="Скільки буде 2+2?")
    ws.cell(row=5, column=2, value="single")
    ws.cell(row=5, column=3, value="3")
    ws.cell(row=5, column=4, value="4")
    ws.cell(row=5, column=5, value="5")
    ws.cell(row=5, column=7, value="2")

    # Row 6: multi-choice example
    ws.cell(row=6, column=1, value="Оберіть голосні букви")
    ws.cell(row=6, column=2, value="multi")
    ws.cell(row=6, column=3, value="А")
    ws.cell(row=6, column=4, value="Е")
    ws.cell(row=6, column=5, value="Б")
    ws.cell(row=6, column=6, value="В")
    ws.cell(row=6, column=7, value="1,2")

    # Row 7: text example
    ws.cell(row=7, column=1, value="Розкажіть про себе")
    ws.cell(row=7, column=2, value="text")

    for col in range(1, 8):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_quiz_xlsx(data: bytes) -> tuple[dict, list[dict], list[str]]:
    import openpyxl

    errors: list[str] = []
    metadata: dict = {}
    questions: list[dict] = []

    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(data), read_only=True)
    except Exception as e:
        return metadata, questions, [f"Не вдалося відкрити файл: {e}"]

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 5:
        return metadata, questions, ["Файл занадто короткий. Використайте шаблон."]

    # Row index 1 (second row) = metadata values
    meta_row = rows[1]
    metadata["title"] = str(meta_row[0]).strip() if meta_row[0] else "Без назви"
    metadata["description"] = str(meta_row[1]).strip() if meta_row[1] else None
    try:
        metadata["time_limit_min"] = int(meta_row[2]) if meta_row[2] else None
    except (ValueError, TypeError):
        metadata["time_limit_min"] = None
    raw_shuffle = str(meta_row[3]).strip().upper() if meta_row[3] else "FALSE"
    metadata["shuffle_questions"] = raw_shuffle in ("TRUE", "1", "YES")

    # Rows 5+ (index 4+) are questions
    for row_idx, row in enumerate(rows[4:], start=5):
        if all(c is None for c in row):
            continue

        q_text = str(row[0]).strip() if row[0] else ""
        q_type = str(row[1]).strip().lower() if row[1] else ""

        if not q_text:
            continue

        if q_type not in ("single", "multi", "text"):
            errors.append(f"Рядок {row_idx}: невідомий тип '{q_type}' (має бути single/multi/text)")
            continue

        options_raw = [str(row[i]).strip() if i < len(row) and row[i] else "" for i in range(2, 6)]
        correct_raw = str(row[6]).strip() if len(row) > 6 and row[6] else ""

        if q_type in ("single", "multi"):
            opts_texts = [o for o in options_raw if o]
            if len(opts_texts) < 2:
                errors.append(f"Рядок {row_idx}: потрібно мінімум 2 варіанти відповіді")
                continue

            try:
                correct_indices = {int(x.strip()) - 1 for x in correct_raw.split(",") if x.strip()}
            except ValueError:
                errors.append(f"Рядок {row_idx}: некоректне поле 'correct' — має бути '1' або '1,2'")
                continue

            if not correct_indices:
                errors.append(f"Рядок {row_idx}: не вказано правильну відповідь")
                continue

            options = [
                {"text": t, "is_correct": i in correct_indices}
                for i, t in enumerate(opts_texts)
            ]
            questions.append({"text": q_text, "type": q_type, "options": options})
        else:
            questions.append({"text": q_text, "type": "text", "options": []})

    if not questions and not errors:
        errors.append("Не знайдено жодного питання. Перевірте формат файлу.")

    return metadata, questions, errors
